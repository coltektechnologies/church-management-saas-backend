"""
Excel export using openpyxl (already in project requirements).
"""

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .base import BaseExporter


class ExcelExporter(BaseExporter):
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    file_extension = "xlsx"

    def export(self, report_data: Dict[str, Any], title: str = "Report") -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"[:31]

        meta = report_data.get("meta") or {}
        row = 1
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        ws.cell(
            row=row,
            column=1,
            value=f"Period: {meta.get('date_from', '')} to {meta.get('date_to', '')}",
        )
        row += 2

        data = report_data.get("data")
        if data is None:
            ws.cell(row=row, column=1, value="No data available.")
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer

        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        if isinstance(data, dict):
            for sheet_name, value in data.items():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    if ws.title == "Report" and row == 3:
                        # Use first list for main sheet
                        self._write_table(ws, value, row, header_fill, header_font)
                        row = ws.max_row + 2
                    else:
                        # New sheet for each other list
                        new_ws = wb.create_sheet(title=str(sheet_name)[:31])
                        self._write_table(new_ws, value, 1, header_fill, header_font)
                elif isinstance(value, (int, str, float)):
                    ws.cell(
                        row=row,
                        column=1,
                        value=str(sheet_name).replace("_", " ").title(),
                    )
                    ws.cell(row=row, column=2, value=value)
                    row += 1
        elif isinstance(data, list) and data and isinstance(data[0], dict):
            self._write_table(ws, data, row, header_fill, header_font)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _write_table(
        self, ws, rows: List[Dict], start_row: int, header_fill, header_font
    ) -> int:
        if not rows:
            return start_row
        headers = list(rows[0].keys())
        for col, h in enumerate(headers, 1):
            c = ws.cell(
                row=start_row, column=col, value=str(h).replace("_", " ").title()
            )
            c.fill = header_fill
            c.font = header_font
        for r_idx, row in enumerate(rows, start_row + 1):
            for c_idx, h in enumerate(headers, 1):
                val = row.get(h, "")
                if hasattr(val, "isoformat"):
                    val = str(val)
                ws.cell(row=r_idx, column=c_idx, value=val)
        return start_row + len(rows) + 1
